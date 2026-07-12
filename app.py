#!/usr/bin/env python3
"""Break Schedule - Al Mana General Hospitals (Server-Sync Version)"""
from flask import Flask, send_from_directory, request, jsonify, send_file
import json, io, os, smtplib, ssl, sqlite3, csv
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from fpdf import FPDF

app = Flask(__name__)
BASE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE, "break_schedule.db")

COL_MAP = {"morning":(2,3),"afternoon":(5,6),"night":(8,9)}
SLB = {"morning":"🌅 Morning (8-5)","afternoon":"☀️ Afternoon (11-8)","night":"🌙 Night (1-10)"}

# ═══════════════ DB SETUP ═══════════════
def get_db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    con = get_db()
    con.executescript("""
        CREATE TABLE IF NOT EXISTS employees(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            email TEXT NOT NULL DEFAULT '',
            code TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS assignments(
            slot_key TEXT PRIMARY KEY,
            employee_name TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS shifts(
            shift TEXT NOT NULL,
            time TEXT NOT NULL,
            idx INTEGER DEFAULT 0,
            PRIMARY KEY(shift,time)
        );
        CREATE TABLE IF NOT EXISTS view_only(
            shift TEXT NOT NULL,
            time TEXT NOT NULL,
            PRIMARY KEY(shift,time)
        );
        CREATE TABLE IF NOT EXISTS custom_durations(
            slot_key TEXT PRIMARY KEY,
            duration INTEGER NOT NULL DEFAULT 15
        );
        CREATE TABLE IF NOT EXISTS recipients(
            name TEXT NOT NULL,
            email TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS duration_pattern(
            id INTEGER PRIMARY KEY DEFAULT 1,
            d0 INTEGER DEFAULT 15,
            d1 INTEGER DEFAULT 15,
            d2 INTEGER DEFAULT 20,
            d3 INTEGER DEFAULT 10
        );
        CREATE TABLE IF NOT EXISTS reports(
            date TEXT PRIMARY KEY,
            data TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS store(
            key TEXT PRIMARY KEY,
            val TEXT
        );
    """)
    # Ensure default duration pattern
    cur = con.execute("SELECT COUNT(*) FROM duration_pattern")
    if cur.fetchone()[0] == 0:
        con.execute("INSERT INTO duration_pattern DEFAULT VALUES")

    # Restore 81 default shifts if table empty
    cur = con.execute("SELECT COUNT(*) FROM shifts")
    if cur.fetchone()[0] == 0:
        defaults = {
            "morning": ["9:00","9:10","9:15","9:20","9:30","9:45","9:50","10:10","10:55","11:10","11:15","11:35","11:40","12:00","12:20","12:40","12:55","13:10","13:25","13:40","14:00","14:20","14:40","14:55","15:00","15:10","15:20","15:30","15:40","15:50"],
            "afternoon": ["12:30","12:40","12:50","13:00","13:10","13:15","13:30","13:40","14:00","14:20","14:50","15:00","15:20","16:00","16:15","16:30","16:40","17:00","17:15","17:35","18:00","18:15","18:30","19:00","19:15","19:35","19:50","20:00","20:10","20:20"],
            "night": ["15:00","15:15","15:30","15:50","16:00","16:20","16:40","17:00","17:20","17:35","18:10","18:45","19:10","19:25","20:20","20:40","20:50","21:10","21:20","21:30","21:40"]
        }
        for shift, times in defaults.items():
            for i, t in enumerate(times):
                con.execute("INSERT OR IGNORE INTO shifts(shift,time,idx) VALUES(?,?,?)", (shift,t,i))
        vo_defaults = {"morning":["12:00","13:10","15:10"],"afternoon":["15:00","18:00","17:00"],"night":["17:00","19:10","16:00"]}
        for shift, times in vo_defaults.items():
            for t in times:
                con.execute("INSERT OR IGNORE INTO view_only(shift,time) VALUES(?,?)", (shift,t))

    # Seed default employees if table empty
    cur = con.execute("SELECT COUNT(*) FROM employees")
    if cur.fetchone()[0] == 0:
        default_emps = [
            ("Afaf Alkhaldi","Afaf.Alkhaldi@almanahospital.com.sa"),
            ("Aisha Jezan","Aisha.Jezan@almanahospital.com.sa"),
            ("Aljawhara Almelifi","Aljawhara.Almelifi@almanahospital.com.sa"),
            ("Aljazi Al-Anazi","Aljazi.Al-Anazi@almanahospital.com.sa"),
            ("Amna Albrari","Amna.Albrari@almanahospital.com.sa"),
            ("Fatimah Busaeed","Fatimah.Busaeed@almanahospital.com.sa"),
            ("Hussain Albaqshi","Hussain.Albaqshi@almanahospital.com.sa"),
            ("Khadra alqahtani","Khadra.alqahtani@almanahospital.com.sa"),
            ("Lina Alshshri","Lina.Alshshri@almanahospital.com.sa"),
            ("Manal Alfarhan","Manal.Alfarhan@almanahospital.com.sa"),
            ("Mariam Almahawsh","Mariam.Almahawsh@almanahospital.com.sa"),
            ("Maryam Alsuwayigh","Maryam.Al-suwayigh@almanahospital.com.sa"),
            ("Nehal Alfarhan","Nehal.Alfarhan@almanahospital.com.sa"),
            ("Rawan Al-Naghmoush","Rawan.Al-Naghmoush@almanahospital.com.sa"),
            ("Rima Alqahtani","Rima.Alqahtani@almanahospital.com.sa"),
            ("Ruqaia AlShammery","Ruqaia.AlShammery@almanahospital.com.sa"),
            ("Ruqaia Zarei","Ruqaia.Zarei@almanahospital.com.sa"),
            ("Saif Alruwaili","Saif.Alruwaili@almanahospital.com.sa"),
            ("Salman Alrhiman","Salman.Alrhiman@almanahospital.com.sa"),
            ("Sara Aldossary","Sara.Aldossary@almanahospital.com.sa"),
            ("Shahad Al-Harbi","Shahad.Al-Harbi@almanahospital.com.sa"),
            ("Sheikah Alkhelaiwi","Sheikah.Alkhelaiwi@almanahospital.com.sa"),
            ("Shouq Al-Harbi","Shouq.Al-Harbi@almanahospital.com.sa"),
            ("Tahani AlSahluli","Tahani.AlSahluli@almanahospital.com.sa"),
            ("Khadijah Almalki","Khadijah.Almalki@almanahospital.com.sa"),
        ]
        for name, email in default_emps:
            con.execute("INSERT OR IGNORE INTO employees(name,email,code) VALUES(?,?,?)", (name,email,""))

    # Seed emp_list and sup_list in store if empty
    row = con.execute("SELECT val FROM store WHERE key='emp_list'").fetchone()
    if not row:
        emp_json = json.dumps([{"name":n,"email":e,"code":""} for n,e in default_emps], ensure_ascii=False)
        con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("emp_list", emp_json))
    row2 = con.execute("SELECT val FROM store WHERE key='sup_list'").fetchone()
    if not row2:
        con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("sup_list", json.dumps(["Shift Supervisor","Head Nurse - Morning","Head Nurse - Afternoon"])))
    con.commit()
    con.close()

init_db()

# ═══════════════ API: DATA ═══════════════
@app.route("/api/data", methods=["GET","POST","OPTIONS"])
def api_data():
    if request.method == "OPTIONS":
        return jsonify({"ok":True})

    con = get_db()

    if request.method == "GET":
        # Load all data
        emps = [{"name":r["name"],"email":r["email"],"code":r["code"]} for r in con.execute("SELECT * FROM employees").fetchall()]
        ass = {r["slot_key"]:r["employee_name"] for r in con.execute("SELECT * FROM assignments").fetchall()}
        sh = {"morning":{"lb":"🌅 Morning (8-5)","ts":[]},"afternoon":{"lb":"☀️ Afternoon (11-8)","ts":[]},"night":{"lb":"🌙 Night (1-10)","ts":[]}}
        for r in con.execute("SELECT * FROM shifts ORDER BY shift, idx"):
            sh[r["shift"]]["ts"].append(r["time"])
        vo = {}
        for r in con.execute("SELECT * FROM view_only"):
            if r["shift"] not in vo: vo[r["shift"]] = []
            vo[r["shift"]].append(r["time"])
        cdur = {r["slot_key"]:r["duration"] for r in con.execute("SELECT * FROM custom_durations").fetchall()}
        rs = [{"name":r["name"],"email":r["email"]} for r in con.execute("SELECT * FROM recipients").fetchall()]
        dp = con.execute("SELECT * FROM duration_pattern WHERE id=1").fetchone()
        dur = [dp["d0"],dp["d1"],dp["d2"],dp["d3"]]
        xb_row = con.execute("SELECT val FROM store WHERE key='extra_breaks'").fetchone()
        xb = json.loads(xb_row["val"]) if xb_row else {}
        xb_req_row = con.execute("SELECT val FROM store WHERE key='extra_breaks_req'").fetchone()
        xb_req = json.loads(xb_req_row["val"]) if xb_req_row else {}

        con.close()
        return jsonify({"employees":emps,"assignments":ass,"shifts":sh,"view_only":vo,"custom_durations":cdur,"recipients":rs,"durations":dur,"extra_breaks":xb,"extra_breaks_req":xb_req})

    else:
        # Save all data
        d = request.json
        if not d: return jsonify({"error":"no data"}), 400

        # ⏰ Auto-save report if assignments are being cleared (daily reset)
        old_ass = {r["slot_key"]:r["employee_name"] for r in con.execute("SELECT * FROM assignments").fetchall()}
        new_ass = d.get("assignments",{})
        if old_ass and not new_ass:
            now = datetime.now()
            date = now.strftime("%Y-%m-%d")
            old_row = con.execute("SELECT data FROM reports WHERE date=?", (date,)).fetchone()
            if not old_row:
                # Build and save a report before clearing
                emps = [{"name":r["name"],"email":r["email"],"code":r["code"]} for r in con.execute("SELECT * FROM employees").fetchall()]
                cdur = {r["slot_key"]:r["duration"] for r in con.execute("SELECT * FROM custom_durations").fetchall()}
                dp = con.execute("SELECT * FROM duration_pattern WHERE id=1").fetchone()
                dur = [dp["d0"],dp["d1"],dp["d2"],dp["d3"]]
                sh = {"morning":{"lb":"🌅 Morning (8-5)"},"afternoon":{"lb":"☀️ Afternoon (11-8)"},"night":{"lb":"🌙 Night (1-10)"}}
                for r in con.execute("SELECT * FROM shifts ORDER BY shift, idx"):
                    if "ts" not in sh[r["shift"]]: sh[r["shift"]]["ts"] = []
                    sh[r["shift"]]["ts"].append(r["time"])
                vo = {}
                for r in con.execute("SELECT * FROM view_only"):
                    if r["shift"] not in vo: vo[r["shift"]] = []
                    vo[r["shift"]].append(r["time"])
                report = {"date":date,"saved_at":now.isoformat(),"employees":emps,"assignments":old_ass,"shifts":sh,"view_only":vo,"custom_durations":cdur,"durations":dur}
                # Include XB (extra breaks) and XB_REQ (requests) in the report
                xb_row = con.execute("SELECT val FROM store WHERE key='extra_breaks'").fetchone()
                if xb_row: report["extra_breaks"] = json.loads(xb_row["val"])
                xb_req_row = con.execute("SELECT val FROM store WHERE key='extra_breaks_req'").fetchone()
                if xb_req_row: report["extra_breaks_req"] = json.loads(xb_req_row["val"])
                con.execute("INSERT OR REPLACE INTO reports(date,data) VALUES(?,?)", (date, json.dumps(report)))

        emps = d.get("employees",[])
        if len(emps) >= 5:  # safety guard: only replace if 5+ employees (prevent accidental wipe)
            con.execute("DELETE FROM employees")
            for e in emps:
                con.execute("INSERT INTO employees(name,email,code) VALUES(?,?,?)", (e["name"],e.get("email",""),e.get("code","")))

        con.execute("DELETE FROM assignments")
        for k,v in d.get("assignments",{}).items():
            con.execute("INSERT INTO assignments(slot_key,employee_name) VALUES(?,?)", (k,v))

        con.execute("DELETE FROM view_only")
        for shift, times in d.get("view_only",{}).items():
            for t in times:
                con.execute("INSERT INTO view_only(shift,time) VALUES(?,?)", (shift,t))

        con.execute("DELETE FROM custom_durations")
        for k,v in d.get("custom_durations",{}).items():
            con.execute("INSERT INTO custom_durations(slot_key,duration) VALUES(?,?)", (k,v))

        con.execute("DELETE FROM recipients") if d.get("recipients") else None
        if d.get("recipients"):
            con.execute("DELETE FROM recipients")
            for r in d["recipients"]:
                con.execute("INSERT INTO recipients(name,email) VALUES(?,?)", (r["name"],r["email"]))

        dur = d.get("durations",[15,15,20,10])
        con.execute("UPDATE duration_pattern SET d0=?,d1=?,d2=?,d3=? WHERE id=1", dur[:4])

        # Sync shifts
        if d.get("shifts"):
            con.execute("DELETE FROM shifts")
            for shift, data in d["shifts"].items():
                for i, t in enumerate(data.get("ts",[])):
                    con.execute("INSERT INTO shifts(shift,time,idx) VALUES(?,?,?)", (shift,t,i))

        # Save extra breaks
        if d.get("extra_breaks") is not None:
            con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("extra_breaks", json.dumps(d["extra_breaks"])))
        if d.get("extra_breaks_req") is not None:
            con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("extra_breaks_req", json.dumps(d["extra_breaks_req"])))

        con.commit()
        con.close()
        return jsonify({"ok":True,"message":"✅ تم الحفظ في السيرفر"})

@app.route("/api/sync", methods=["POST"])
def api_sync():
    """Quick sync - just assignments + check reset"""
    d = request.json
    con = get_db()
    if d.get("assignments"):
        con.execute("DELETE FROM assignments")
        for k,v in d["assignments"].items():
            con.execute("INSERT OR REPLACE INTO assignments(slot_key,employee_name) VALUES(?,?)", (k,v))
    con.commit()
    con.close()
    return jsonify({"ok":True})



# ═══════════════ BADGE CODES (تخزين منفصل عن التصفير) ═══════════════
@app.route("/api/badges", methods=["GET","POST"])
def api_badges():
    """Get or set badge codes independently — never cleared by daily reset"""
    con = get_db()
    if request.method == "GET":
        codes = {r["name"]:r["code"] for r in con.execute("SELECT name, code FROM employees WHERE code != ''").fetchall()}
        con.close()
        return jsonify(codes)
    else:
        d = request.json
        if not d: return jsonify({"error":"no data"}), 400
        for name, code in d.items():
            con.execute("UPDATE employees SET code=? WHERE name=?", (str(code), name))
        con.commit()
        con.close()
        return jsonify({"ok":True, "message":"✅ تم حفظ البادجات"})
# ═══════════════ REPORTS (حفظ تقارير يومية) ═══════════════

@app.route("/api/emp-list", methods=["GET","POST"])
def api_emp_list():
    """Get/set employee list independently — not cleared by daily reset"""
    con = get_db()
    if request.method == "GET":
        # Return from employees table (authoritative source — always has full list)
        rows = con.execute("SELECT name, email, code FROM employees ORDER BY id").fetchall()
        emp_list = [{"name":r["name"],"email":r["email"],"code":r["code"]} for r in rows]
        con.close()
        return jsonify(emp_list)
    else:
        d = request.json
        if d is None: return jsonify({"error":"no data"}), 400
        if isinstance(d, list) and len(d) > 3:  # safety: ignore empty/damaged payloads
            con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("emp_list", json.dumps(d)))
        elif isinstance(d, list) and len(d) <= 3:
            return jsonify({"ok":True, "warning":"list too small — keeping existing"})
        con.commit()
        con.close()
        return jsonify({"ok":True, "message":"✅ تم حفظ الموظفين"})



# ═══════════════ BREAK TIMER (عداد البريك المباشر) ═══════════════
@app.route("/api/sup-list", methods=["GET","POST"])
def api_sup_list():
    """Get/set supervisor list independently — not cleared by daily reset"""
    con = get_db()
    if request.method == "GET":
        row = con.execute("SELECT val FROM store WHERE key='sup_list'").fetchone()
        con.close()
        return jsonify(json.loads(row["val"]) if row else [])
    else:
        d = request.json
        if d is None: return jsonify({"error":"no data"}), 400
        if isinstance(d, list) and len(d) > 0:
            con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("sup_list", json.dumps(d)))
        elif isinstance(d, list) and len(d) == 0:
            return jsonify({"ok":True, "warning":"empty list ignored — keeping existing"})
        con.commit()
        con.close()
        return jsonify({"ok":True, "message":"✅ تم حفظ المشرفين"})


# ═══════════════ EXTRA BREAKS (البريك الإضافي) ═══════════════
@app.route("/api/x-breaks", methods=["GET","POST"])
def api_x_breaks():
    """Get/set extra breaks — supervisor grants additional break slots"""
    con = get_db()
    if request.method == "GET":
        row = con.execute("SELECT val FROM store WHERE key='extra_breaks'").fetchone()
        con.close()
        return jsonify(json.loads(row["val"]) if row else {})
    else:
        d = request.json
        if d is None: return jsonify({"error":"no data"}), 400
        con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("extra_breaks", json.dumps(d)))
        con.commit()
        con.close()
        return jsonify({"ok":True, "message":"✅ تم حفظ البريكات الإضافية"})

# ═══════════════ EXTRA BREAKS REQUESTS (طلبات البريك الإضافي) ═══════════════
@app.route("/api/x-breaks-req", methods=["GET","POST"])
def api_x_breaks_req():
    """Get/set extra break requests — employee → supervisor"""
    con = get_db()
    if request.method == "GET":
        row = con.execute("SELECT val FROM store WHERE key='extra_breaks_req'").fetchone()
        con.close()
        return jsonify(json.loads(row["val"]) if row else {})
    else:
        d = request.json
        if d is None: return jsonify({"error":"no data"}), 400
        con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("extra_breaks_req", json.dumps(d)))
        con.commit()
        con.close()
        return jsonify({"ok":True, "message":"✅ تم حفظ طلبات البريكات"})


# ═══════════════📊 EMPLOYEE STATISTICS (الإحصائيات) ═══════════════
@app.route("/api/stats", methods=["POST"])
def api_stats():
    """Compute per-employee stats from saved reports within a date range"""
    d = request.json or {}
    fr = d.get("from","")
    to = d.get("to","")
    con = get_db()

    # Get all reports in range
    if fr and to:
        rows = con.execute("SELECT date, data FROM reports WHERE date >= ? AND date <= ? ORDER BY date", (fr, to)).fetchall()
    else:
        # Last 30 days default
        rows = con.execute("SELECT date, data FROM reports ORDER BY date DESC LIMIT 31").fetchall()

    con.close()

    if not rows:
        return jsonify({"stats":[],"summary":{"total_breaks":0,"total_minutes":0,"total_xb_req":0,"total_xb_app":0,"days":0,"from":"","to":""}})

    # Employee info — collect from all reports
    emp_info = {}
    merged = []
    days = 0
    for row in rows:
        days += 1
        report = json.loads(row["data"])
        merged.append(report)
        for e in report.get("employees",[]):
            emp_info[e["name"]] = {"email":e.get("email",""),"code":e.get("code","")}

    by_emp = {}
    for report in merged:
        date = report.get("date","")
        assigns = report.get("assignments",{})
        shifts = report.get("shifts",{})
        cdur = report.get("custom_durations",{})
        dur = report.get("durations",[15,15,20,10])
        xb_req = report.get("extra_breaks_req",{})
        xb_app = report.get("extra_breaks",{})

        # Per-slot durations map
        slot_dur = {}
        for k in assigns:
            parts = k.split("_")
            if len(parts) != 2: continue
            shift, t = parts
            shift_times = shifts.get(shift,{}).get("ts",[])
            try:
                idx = shift_times.index(t)
            except ValueError:
                idx = 0
            slot_dur[k] = cdur.get(k) or dur[idx % 4]

        for emp_name in list(set(list(assigns.values()) + list(xb_req.keys()) + list(xb_app.keys()))):
            if emp_name not in by_emp:
                by_emp[emp_name] = {"breaks":0,"minutes":0,"xb_req":0,"xb_app":0,"daily":{},"slots":[]}

            emp = by_emp[emp_name]
            if date not in emp["daily"]:
                emp["daily"][date] = {"breaks":0,"minutes":0,"slots":[],"xb_req":0,"xb_app":0}

            # Count breaks
            for k, v in assigns.items():
                if v == emp_name:
                    emp["breaks"] += 1
                    emp["minutes"] += slot_dur.get(k, 15)
                    emp["slots"].append(k)
                    emp["daily"][date]["breaks"] += 1
                    emp["daily"][date]["minutes"] += slot_dur.get(k, 15)
                    emp["daily"][date]["slots"].append(k)

            # Extra breaks
            if emp_name in xb_req:
                emp["xb_req"] += 1
                emp["daily"][date]["xb_req"] += 1
            if emp_name in xb_app:
                emp["xb_app"] += 1
                emp["daily"][date]["xb_app"] += 1

    # Build result — sort by most breaks
    def_tm = lambda k: "15"
    stats = []
    for name, data in sorted(by_emp.items(), key=lambda x: -x[1]["breaks"]):
        # Find most common slot
        slot_counts = {}
        for s in data["slots"]:
            slot_counts[s] = slot_counts.get(s, 0) + 1
        top_slot = max(slot_counts, key=slot_counts.get) if slot_counts else ""
        # Clean up slot name
        if top_slot:
            parts = top_slot.split("_",1)
            top_slot = parts[1] if len(parts) > 1 else top_slot

        emp = emp_info.get(name,{})
        stats.append({
            "name": name,
            "email": emp.get("email",""),
            "breaks": data["breaks"],
            "minutes": data["minutes"],
            "xb_req": data["xb_req"],
            "xb_app": data["xb_app"],
            "avg_per_day": round(data["breaks"] / days, 1) if days else 0,
            "top_slot": top_slot,
            "daily": {d:data["daily"][d] for d in sorted(data["daily"].keys())}
        })

    total = sum(s["breaks"] for s in stats)
    total_min = sum(s["minutes"] for s in stats)
    total_xb_r = sum(s["xb_req"] for s in stats)
    total_xb_a = sum(s["xb_app"] for s in stats)

    return jsonify({
        "stats": stats,
        "summary": {
            "total_breaks": total,
            "total_minutes": total_min,
            "total_xb_req": total_xb_r,
            "total_xb_app": total_xb_a,
            "employees": len(stats),
            "days": days,
            "from": merged[-1]["date"] if merged else "",
            "to": merged[0]["date"] if merged else ""
        }
    })


# ═══════════════📄 PDF REPORT (تقرير PDF شهري) ═══════════════
@app.route("/api/report/pdf", methods=["POST"])
def api_report_pdf():
    """Generate a PDF report with employee statistics"""
    d = request.json or {}
    fr = d.get("from","")
    to = d.get("to","")
    con = get_db()

    if fr and to:
        rows = con.execute("SELECT date, data FROM reports WHERE date >= ? AND date <= ? ORDER BY date", (fr, to)).fetchall()
    else:
        rows = con.execute("SELECT date, data FROM reports ORDER BY date DESC LIMIT 31").fetchall()
    con.close()

    if not rows:
        return jsonify({"error":"لا توجد تقارير في هذا النطاق"}), 404

    # Parse all reports
    merged = []
    for row in rows:
        report = json.loads(row["data"])
        merged.append(report)

    emp_info = {}
    for r in merged:
        for e in r.get("employees",[]):
            if e["name"] not in emp_info:
                emp_info[e["name"]] = True

    # Compute stats (same logic as /api/stats)
    by_emp = {}
    for report in merged:
        date = report.get("date","")
        assigns = report.get("assignments",{})
        shifts = report.get("shifts",{})
        cdur = report.get("custom_durations",{})
        dur = report.get("durations",[15,15,20,10])
        xb_req = report.get("extra_breaks_req",{})
        xb_app = report.get("extra_breaks",{})

        slot_dur = {}
        for k in assigns:
            parts = k.split("_",1)
            if len(parts) != 2: continue
            shift_t = parts[0]
            t = parts[1]
            shift_times = shifts.get(shift_t,{}).get("ts",[])
            try:
                idx = shift_times.index(t)
            except ValueError:
                idx = 0
            slot_dur[k] = cdur.get(k) or dur[idx % 4]

        all_names = list(set(list(assigns.values())))
        for emp_name in all_names:
            if emp_name not in by_emp:
                by_emp[emp_name] = {"breaks":0,"minutes":0,"xb_req":0,"xb_app":0}
            if emp_name in xb_req: by_emp[emp_name]["xb_req"] += 1
            if emp_name in xb_app: by_emp[emp_name]["xb_app"] += 1
            for k, v in assigns.items():
                if v == emp_name:
                    by_emp[emp_name]["breaks"] += 1
                    by_emp[emp_name]["minutes"] += slot_dur.get(k, 15)

    sorted_emps = sorted(by_emp.items(), key=lambda x: -x[1]["breaks"])

    # ── Build PDF ──
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_font("Noto", "", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", uni=True)
    pdf.add_page()

    # Title
    pdf.set_font("Noto", "", 18)
    pdf.cell(0, 12, "مستشفيات المانع العامة", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Noto", "", 12)
    fr_label = fr or merged[-1]["date"]
    to_label = to or merged[0]["date"]
    pdf.cell(0, 8, f"تقرير البريكات الشامل", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Noto", "", 10)
    pdf.cell(0, 7, f"{fr_label}  ~  {to_label}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(3)

    # Summary box
    total_b = sum(e[1]["breaks"] for e in sorted_emps)
    total_m = sum(e[1]["minutes"] for e in sorted_emps)
    total_x = sum(e[1]["xb_req"] for e in sorted_emps)
    total_xa = sum(e[1]["xb_app"] for e in sorted_emps)
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("Noto", "", 10)
    summary_text = (
        f"إجمالي البريكات: {total_b}   |   إجمالي الدقائق: {total_m}   |   "
        f"طلبات إضافية: {total_x}   |   موافقات: {total_xa}   |   "
        f"عدد الموظفين: {len(sorted_emps)}   |   الأيام: {len(rows)}"
    )
    pdf.cell(0, 8, summary_text, new_x="LMARGIN", new_y="NEXT", align="C", fill=True)
    pdf.ln(5)

    # ── Employee Table ──
    pdf.set_font("Noto", "", 9)
    col_w = [46, 18, 18, 18, 18, 18, 18, 18]
    headers = ["الموظف", "بريكات", "دقائق", "م/يوم", "طلب", "موافقة", ""]
    pdf.set_fill_color(41, 128, 185)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    for idx, (name, data) in enumerate(sorted_emps):
        if idx % 2 == 0:
            pdf.set_fill_color(245, 248, 252)
        else:
            pdf.set_fill_color(255, 255, 255)
        avg = round(data["breaks"] / len(rows), 1)
        vals = [name, str(data["breaks"]), str(data["minutes"]), str(avg), str(data["xb_req"]), str(data["xb_app"]), ""]
        for i, v in enumerate(vals):
            pdf.cell(col_w[i], 6, v, border=1, fill=True, align="C")
        pdf.ln()

        # Page break if needed
        if pdf.get_y() > 260:
            pdf.add_page()
            pdf.set_font("Noto", "", 9)
            pdf.set_fill_color(41, 128, 185)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(headers):
                pdf.cell(col_w[i], 7, h, border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_text_color(0, 0, 0)

    # ── Daily Breakdown ──
    pdf.ln(8)
    pdf.set_font("Noto", "", 13)
    pdf.set_text_color(41, 128, 185)
    pdf.cell(0, 10, "التفصيل اليومي", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_text_color(0, 0, 0)

    for report in merged:
        date = report.get("date","")
        assigns = report.get("assignments",{})
        shifts = report.get("shifts",{})
        cdur = report.get("custom_durations",{})
        dur = report.get("durations",[15,15,20,10])
        xb_req = report.get("extra_breaks_req",{})
        xb_app = report.get("extra_breaks",{})

        pdf.ln(3)
        pdf.set_fill_color(230, 240, 250)
        pdf.set_font("Noto", "", 11)
        day_breaks = len(assigns)
        day_xb = len(xb_req) + len(xb_app)
        pdf.cell(0, 7, f"{date}  —  {day_breaks} بريك", new_x="LMARGIN", new_y="NEXT", align="C", fill=True)
        pdf.set_font("Noto", "", 8)

        # Group by shift
        by_shift = {"morning":[],"afternoon":[],"night":[]}
        for k, v in assigns.items():
            parts = k.split("_",1)
            shift_name = parts[0] if len(parts) == 2 else ""
            t = parts[1] if len(parts) == 2 else k
            dr = cdur.get(k) or 15
            if shift_name in by_shift:
                by_shift[shift_name].append((t, v, dr))

        for sname, sdata in by_shift.items():
            if not sdata: continue
            slb = {"morning":"🌅 صباحي","afternoon":"☀️ مسائي","night":"🌙 ليلي"}
            sdata.sort(key=lambda x: x[0])
            times = ", ".join(f"{t} ({v}-{dr}د)" for t, v, dr in sdata)
            pdf.cell(0, 5, f"  {slb.get(sname,sname)}: {times}", new_x="LMARGIN", new_y="NEXT")

        if xb_req or xb_app:
            xb_line = ""
            if xb_req: xb_line += f"طلبات: {', '.join(xb_req.keys())}"
            if xb_app: xb_line += f"  |  موافقات: {', '.join(xb_app.keys())}"
            pdf.set_font("Noto", "", 8)
            pdf.cell(0, 5, f"  📩 {xb_line}", new_x="LMARGIN", new_y="NEXT")

        # Page break
        if pdf.get_y() > 270:
            pdf.add_page()

    # Footer
    pdf.ln(10)
    pdf.set_font("Noto", "", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, f"Al Mana General Hospitals  |  مستشفيات المانع العامة  |  {datetime.now().strftime('%Y-%m-%d %I:%M %p')}", new_x="LMARGIN", new_y="NEXT", align="C")

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)

    fn = f"Break_Report_{fr or 'all'}_{to or 'all'}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fn)
    """Get/set break timer data — tracks active breaks per employee per day"""
    con = get_db()
    if request.method == "GET":
        row = con.execute("SELECT val FROM store WHERE key='break_timer'").fetchone()
        con.close()
        return jsonify(json.loads(row["val"]) if row else {})
    else:
        d = request.json
        if d is None: return jsonify({"error":"no data"}), 400
        con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", ("break_timer", json.dumps(d)))
        con.commit()
        con.close()
        return jsonify({"ok":True})

@app.route("/api/report/save", methods=["POST"])
def api_save_report():
    """Save current assignments as a daily report (date-based)"""
    con = get_db()
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")

    emps = [{"name":r["name"],"email":r["email"],"code":r["code"]} for r in con.execute("SELECT * FROM employees").fetchall()]
    ass = {r["slot_key"]:r["employee_name"] for r in con.execute("SELECT * FROM assignments").fetchall()}
    cdur = {r["slot_key"]:r["duration"] for r in con.execute("SELECT * FROM custom_durations").fetchall()}
    dp = con.execute("SELECT * FROM duration_pattern WHERE id=1").fetchone()
    dur = [dp["d0"],dp["d1"],dp["d2"],dp["d3"]]
    sh = {"morning":{"lb":"🌅 Morning (8-5)"},"afternoon":{"lb":"☀️ Afternoon (11-8)"},"night":{"lb":"🌙 Night (1-10)"}}
    for r in con.execute("SELECT * FROM shifts ORDER BY shift, idx"):
        if "ts" not in sh[r["shift"]]: sh[r["shift"]]["ts"] = []
        sh[r["shift"]]["ts"].append(r["time"])
    vo = {}
    for r in con.execute("SELECT * FROM view_only"):
        if r["shift"] not in vo: vo[r["shift"]] = []
        vo[r["shift"]].append(r["time"])

    report = {
        "date": date,
        "saved_at": now.isoformat(),
        "employees": emps,
        "assignments": ass,
        "shifts": sh,
        "view_only": vo,
        "custom_durations": cdur,
        "durations": dur,
        "break_timer": json.loads(con.execute("SELECT val FROM store WHERE key='break_timer'").fetchone()["val"]) if con.execute("SELECT val FROM store WHERE key='break_timer'").fetchone() else {}
    }

    con.execute("INSERT OR REPLACE INTO reports(date,data) VALUES(?,?)", (date, json.dumps(report)))
    con.commit()
    count_breaks = len(ass)
    con.close()
    return jsonify({"ok":True, "date":date, "breaks":count_breaks, "message":f"✅ تم حفظ تقرير {date}"})

@app.route("/api/report/list", methods=["GET"])
def api_list_reports():
    con = get_db()
    rows = con.execute("SELECT date FROM reports ORDER BY date DESC").fetchall()
    con.close()
    return jsonify([r["date"] for r in rows])

@app.route("/api/report/get/<date>", methods=["GET"])
def api_get_report(date):
    con = get_db()
    row = con.execute("SELECT data FROM reports WHERE date = ?", (date,)).fetchone()
    con.close()
    if not row:
        return jsonify({"error":"not found"}), 404
    return jsonify(json.loads(row["data"]))

# ═══════════════ STATIC FILES ═══════════════
@app.route("/")
@app.route("/<path:filename>")
def idx(filename="schedule.html"):
    return send_from_directory(BASE, filename)

# ═══════════════ EXCEL EXPORT ═══════════════
@app.route("/api/export-excel", methods=["POST"])
def export():
    try:
        d = request.json
        assigns = d.get("assignments",{})
        shifts = d.get("shifts",{})
        vo = d.get("view_only",{})
        dur = d.get("durations",[15,15,20,10])
        cdur = d.get("custom_durations",{})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "breaksheet"

        gf = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        rf = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        gy = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        bf = Font(name="Calibri", bold=True, size=10)
        nf = Font(name="Calibri", size=10)
        vf = Font(name="Calibri", size=10, italic=True, color="7F8C8D")
        hf = Font(name="Calibri", bold=True, size=14, color="1A2A3A")

        ws["A1"] = f"Break Schedule - Al Mana General Hospitals  |  {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"
        ws["A1"].font = hf

        row = 3
        for sk, (tc, ec) in COL_MAP.items():
            ts = shifts.get(sk,{}).get("ts",[])
            if not ts: continue
            ws.cell(row=row, column=2, value=SLB.get(sk,sk)).font = Font(name="Calibri", bold=True, size=11)
            row += 1
            for i, t in enumerate(ts):
                h,m = map(int,t.split(":"))
                k = f"{sk}_{t}"
                vw = t in vo.get(sk,[])
                emp = assigns.get(k)
                dr = cdur.get(k) or dur[i%4]
                ws.cell(row=row, column=tc, value=f"{h%12 or 12}:{m:02d} {'PM' if h>=12 else 'AM'}").font = Font(name="Calibri", bold=True, size=10)
                cell = ws.cell(row=row, column=ec)
                if vw: cell.value,cell.fill,cell.font = "📖 View Only",gy,vf
                elif emp: cell.value,cell.fill,cell.font = f"{emp} ({dr} min)",rf,bf
                else: cell.value,cell.fill,cell.font = "🟢 Available",gf,nf
                row += 1
            row += 1

        ws.column_dimensions["A"].width = 2
        for c in ["B","E","H"]: ws.column_dimensions[c].width = 14
        for c in ["C","F","I"]: ws.column_dimensions[c].width = 26

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"Break_Schedule_{datetime.now().strftime('%Y%m%d')}.xlsx")
    except Exception as e:
        return jsonify({"error":str(e)}), 500

# ═══════════════ FULL DATA EXPORT (CSV) ═══════════════
@app.route("/api/export/csv", methods=["GET"])
def api_export_csv():
    """Export ALL data as multi-section CSV file download"""
    con = get_db()

    emps = [{"name":r["name"],"email":r["email"],"code":r["code"]} for r in con.execute("SELECT * FROM employees").fetchall()]
    ass = {r["slot_key"]:r["employee_name"] for r in con.execute("SELECT * FROM assignments").fetchall()}
    sh_rows = [{"shift":r["shift"],"time":r["time"]} for r in con.execute("SELECT * FROM shifts ORDER BY shift, idx").fetchall()]
    vo_rows = [{"shift":r["shift"],"time":r["time"]} for r in con.execute("SELECT * FROM view_only").fetchall()]
    bt_row = con.execute("SELECT val FROM store WHERE key='break_timer'").fetchone()
    bt_val = json.loads(bt_row["val"]) if bt_row else {}
    sl_row = con.execute("SELECT val FROM store WHERE key='sup_list'").fetchone()
    sl_val = json.loads(sl_row["val"]) if sl_row else []
    dp = con.execute("SELECT * FROM duration_pattern WHERE id=1").fetchone()
    cd_rows = [{"slot_key":r["slot_key"],"duration":r["duration"]} for r in con.execute("SELECT * FROM custom_durations").fetchall()]

    con.close()

    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["# Employees", "name", "email", "badge_code"])
    for e in emps:
        w.writerow(["", e["name"], e["email"], e["code"]])

    w.writerow([])
    w.writerow(["# Assignments", "slot_key", "employee_name"])
    for k, v in ass.items():
        w.writerow(["", k, v])

    w.writerow([])
    w.writerow(["# Shifts", "shift", "time"])
    for r in sh_rows:
        w.writerow(["", r["shift"], r["time"]])

    w.writerow([])
    w.writerow(["# View Only", "shift", "time"])
    for r in vo_rows:
        w.writerow(["", r["shift"], r["time"]])

    w.writerow([])
    w.writerow(["# Break Timer", "key", "value"])
    w.writerow(["", "break_timer", json.dumps(bt_val, ensure_ascii=False)])

    w.writerow([])
    w.writerow(["# Supervisor List", "key", "value"])
    w.writerow(["", "sup_list", json.dumps(sl_val, ensure_ascii=False)])

    w.writerow([])
    w.writerow(["# Duration Pattern", "id", "d0", "d1", "d2", "d3"])
    if dp:
        w.writerow(["", dp["id"], dp["d0"], dp["d1"], dp["d2"], dp["d3"]])

    w.writerow([])
    w.writerow(["# Custom Durations", "slot_key", "duration"])
    for r in cd_rows:
        w.writerow(["", r["slot_key"], r["duration"]])

    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    return send_file(out, mimetype="text/csv; charset=utf-8",
                     as_attachment=True,
                     download_name=f"BreakSchedule_Full_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")

# ═══════════════ FULL DATA EXPORT (XLSX) ═══════════════
@app.route("/api/export/xlsx", methods=["GET"])
def api_export_xlsx():
    """Export ALL data as multi-sheet Excel file download"""
    con = get_db()

    emps = [{"name":r["name"],"email":r["email"],"code":r["code"]} for r in con.execute("SELECT * FROM employees").fetchall()]
    ass = {r["slot_key"]:r["employee_name"] for r in con.execute("SELECT * FROM assignments").fetchall()}
    sh_rows = [{"shift":r["shift"],"time":r["time"]} for r in con.execute("SELECT * FROM shifts ORDER BY shift, idx").fetchall()]
    vo_rows = [{"shift":r["shift"],"time":r["time"]} for r in con.execute("SELECT * FROM view_only").fetchall()]
    bt_row = con.execute("SELECT val FROM store WHERE key='break_timer'").fetchone()
    bt_val = json.loads(bt_row["val"]) if bt_row else {}
    sl_row = con.execute("SELECT val FROM store WHERE key='sup_list'").fetchone()
    sl_val = json.loads(sl_row["val"]) if sl_row else []
    dp = con.execute("SELECT * FROM duration_pattern WHERE id=1").fetchone()
    cd_rows = [{"slot_key":r["slot_key"],"duration":r["duration"]} for r in con.execute("SELECT * FROM custom_durations").fetchall()]

    con.close()

    wb = openpyxl.Workbook()

    # 1 — Employees
    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["name", "email", "badge_code"])
    for e in emps:
        ws1.append([e["name"], e["email"], e["code"]])

    # 2 — Assignments
    ws2 = wb.create_sheet("Assignments")
    ws2.append(["slot_key", "employee_name"])
    for k, v in ass.items():
        ws2.append([k, v])

    # 3 — Shifts
    ws3 = wb.create_sheet("Shifts")
    ws3.append(["shift", "time"])
    for r in sh_rows:
        ws3.append([r["shift"], r["time"]])

    # 4 — View Only
    ws4 = wb.create_sheet("View Only")
    ws4.append(["shift", "time"])
    for r in vo_rows:
        ws4.append([r["shift"], r["time"]])

    # 5 — Break Timer
    ws5 = wb.create_sheet("Break Timer")
    ws5.append(["key", "value"])
    ws5.append(["break_timer", json.dumps(bt_val, ensure_ascii=False)])

    # 6 — Supervisor List
    ws6 = wb.create_sheet("Supervisor List")
    ws6.append(["key", "value"])
    ws6.append(["sup_list", json.dumps(sl_val, ensure_ascii=False)])

    # 7 — Duration Pattern
    ws7 = wb.create_sheet("Duration Pattern")
    ws7.append(["id", "d0", "d1", "d2", "d3"])
    if dp:
        ws7.append([dp["id"], dp["d0"], dp["d1"], dp["d2"], dp["d3"]])

    # 8 — Custom Durations
    ws8 = wb.create_sheet("Custom Durations")
    ws8.append(["slot_key", "duration"])
    for r in cd_rows:
        ws8.append([r["slot_key"], r["duration"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"BreakSchedule_Full_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

# ═══════════════ AUTO-BACKUP (triggered from JS every 10 saves) ═══════════════
@app.route("/api/export/backup", methods=["POST"])
def api_export_backup():
    """Generate a timestamped XLSX backup file on the server"""
    BACKUP_DIR = os.path.join(BASE, "backups")
    os.makedirs(BACKUP_DIR, exist_ok=True)
    now = datetime.now()
    stamp = now.strftime("%Y%m%d_%H%M%S")
    fname = f"BreakSchedule_Backup_{stamp}.xlsx"
    fpath = os.path.join(BACKUP_DIR, fname)

    con = get_db()
    emps = [{"name":r["name"],"email":r["email"],"code":r["code"]} for r in con.execute("SELECT * FROM employees").fetchall()]
    ass = {r["slot_key"]:r["employee_name"] for r in con.execute("SELECT * FROM assignments").fetchall()}
    sh_rows = [{"shift":r["shift"],"time":r["time"]} for r in con.execute("SELECT * FROM shifts ORDER BY shift, idx").fetchall()]
    vo_rows = [{"shift":r["shift"],"time":r["time"]} for r in con.execute("SELECT * FROM view_only").fetchall()]
    bt_row = con.execute("SELECT val FROM store WHERE key='break_timer'").fetchone()
    bt_val = json.loads(bt_row["val"]) if bt_row else {}
    sl_row = con.execute("SELECT val FROM store WHERE key='sup_list'").fetchone()
    sl_val = json.loads(sl_row["val"]) if sl_row else []
    dp = con.execute("SELECT * FROM duration_pattern WHERE id=1").fetchone()
    cd_rows = [{"slot_key":r["slot_key"],"duration":r["duration"]} for r in con.execute("SELECT * FROM custom_durations").fetchall()]
    con.close()

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Employees"
    ws1.append(["name","email","badge_code"])
    for e in emps: ws1.append([e["name"],e["email"],e["code"]])
    ws2 = wb.create_sheet("Assignments")
    ws2.append(["slot_key","employee_name"])
    for k,v in ass.items(): ws2.append([k,v])
    ws3 = wb.create_sheet("Shifts")
    ws3.append(["shift","time"])
    for r in sh_rows: ws3.append([r["shift"],r["time"]])
    ws4 = wb.create_sheet("View Only")
    ws4.append(["shift","time"])
    for r in vo_rows: ws4.append([r["shift"],r["time"]])
    ws5 = wb.create_sheet("Break Timer")
    ws5.append(["key","value"]); ws5.append(["break_timer",json.dumps(bt_val,ensure_ascii=False)])
    ws6 = wb.create_sheet("Supervisor List")
    ws6.append(["key","value"]); ws6.append(["sup_list",json.dumps(sl_val,ensure_ascii=False)])
    ws7 = wb.create_sheet("Duration Pattern")
    ws7.append(["id","d0","d1","d2","d3"])
    if dp: ws7.append([dp["id"],dp["d0"],dp["d1"],dp["d2"],dp["d3"]])
    ws8 = wb.create_sheet("Custom Durations")
    ws8.append(["slot_key","duration"])
    for r in cd_rows: ws8.append([r["slot_key"],r["duration"]])
    wb.save(fpath)

    return jsonify({"ok":True,"file":fname,"path":fpath,"message":f"✅ تم حفظ النسخة الاحتياطية: {fname}"})

# ═══════════════ BACKUP Download ═══════════════
@app.route("/api/backup/download", methods=["GET"])
def api_backup_download():
    """Download full data as JSON"""
    import json, io
    con = get_db()
    
    emps = [{"id":r["id"],"name":r["name"],"email":r["email"],"code":r["code"]} for r in con.execute("SELECT * FROM employees ORDER BY id").fetchall()]
    ass = {r["slot_key"]:r["employee_name"] for r in con.execute("SELECT * FROM assignments").fetchall()}
    store = {}
    for r in con.execute("SELECT * FROM store").fetchall():
        try: store[r["key"]] = json.loads(r["val"])
        except: store[r["key"]] = r["val"]
    sh = [{"shift":r["shift"],"time":r["time"],"idx":r["idx"]} for r in con.execute("SELECT * FROM shifts ORDER BY shift, idx").fetchall()]
    vo = [{"shift":r["shift"],"time":r["time"]} for r in con.execute("SELECT * FROM view_only").fetchall()]
    cd = [{"slot_key":r["slot_key"],"duration":r["duration"]} for r in con.execute("SELECT * FROM custom_durations").fetchall()]
    rcpts = [{"name":r["name"],"email":r["email"]} for r in con.execute("SELECT * FROM recipients").fetchall()]
    dp = dict(con.execute("SELECT * FROM duration_pattern WHERE id=1").fetchone() or {})
    
    con.close()
    
    data = {
        "backup_date": datetime.now().isoformat(),
        "version": "2.0",
        "employees": emps,
        "assignments": ass,
        "store": store,
        "shifts": sh,
        "view_only": vo,
        "custom_durations": cd,
        "recipients": rcpts,
        "duration_pattern": dp
    }
    
    buf = io.StringIO()
    json.dump(data, buf, ensure_ascii=False, indent=2)
    out = io.BytesIO(buf.getvalue().encode("utf-8"))
    return send_file(out, mimetype="application/json", as_attachment=True,
                     download_name=f"BreakSchedule_Backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json")

# ═══════════════ BACKUP Restore ═══════════════
@app.route("/api/backup/restore", methods=["POST"])
def api_backup_restore():
    """Restore full data from uploaded JSON backup"""
    try:
        d = request.json
        if not d or not d.get("employees"):
            return jsonify({"error":"⚠️ ملف باك أب غير صحيح"}), 400
        
        con = get_db()
        
        # Restore employees
        if d.get("employees"):
            con.execute("DELETE FROM employees")
            for e in d["employees"]:
                con.execute("INSERT INTO employees(name,email,code) VALUES(?,?,?)",
                          (e["name"], e.get("email",""), e.get("code","")))
        
        # Restore assignments
        con.execute("DELETE FROM assignments")
        for k, v in d.get("assignments",{}).items():
            con.execute("INSERT INTO assignments(slot_key,employee_name) VALUES(?,?)", (k, v))
        
        # Restore store (emp_list, sup_list, break_timer)
        for k, v in d.get("store",{}).items():
            con.execute("INSERT OR REPLACE INTO store(key,val) VALUES(?,?)", (k, json.dumps(v, ensure_ascii=False)))
        
        # Restore shifts
        con.execute("DELETE FROM shifts")
        for s in d.get("shifts",[]):
            con.execute("INSERT INTO shifts(shift,time,idx) VALUES(?,?,?)", (s["shift"], s["time"], s.get("idx",0)))
        
        # Restore view_only
        con.execute("DELETE FROM view_only")
        for v in d.get("view_only",[]):
            con.execute("INSERT INTO view_only(shift,time) VALUES(?,?)", (v["shift"], v["time"]))
        
        # Restore custom_durations
        con.execute("DELETE FROM custom_durations")
        for c in d.get("custom_durations",[]):
            con.execute("INSERT INTO custom_durations(slot_key,duration) VALUES(?,?)", (c["slot_key"], c["duration"]))
        
        con.commit()
        con.close()
        
        return jsonify({"ok":True, "message":f"✅ تم استعادة {len(d.get('employees',[]))} موظف"})
    except Exception as e:
        return jsonify({"error":f"❌ فشل الاستعادة: {str(e)}"}), 500

# ═══════════════ EMAIL ═══════════════
def build_email_body(assigns, shifts, vo, dur, cdur, emps, extras):
    slb = {"morning":"🌅 Morning (8AM-5PM)","afternoon":"☀️ Afternoon (11AM-8PM)","night":"🌙 Night (1PM-10PM)"}
    lines = ["="*55,
             f"  BREAK SCHEDULE — {datetime.now().strftime('%A, %B %d, %Y')}",
             "  Al Mana General Hospitals - مستشفيات المانع العامة",
             "="*55,""]
    for sk, lb in slb.items():
        lines.append(f"── {lb} ──")
        ts = shifts.get(sk, {}).get("ts", [])
        for i, t in enumerate(ts):
            h,m = map(int,t.split(":"))
            ft = f"{h%12 or 12}:{m:02d} {'PM' if h>=12 else 'AM'}"
            k = f"{sk}_{t}"; vw = t in vo.get(sk, []); dr = cdur.get(k) or dur[i%4]; emp = assigns.get(k)
            if emp: lines.append(f"  {ft.ljust(12)}│{str(dr)+'m'.rjust(4)}│{emp}")
            elif vw: lines.append(f"  {ft.ljust(12)}│{str(dr)+'m'.rjust(4)}│📖 مراجعة")
            else: lines.append(f"  {ft.ljust(12)}│{str(dr)+'m'.rjust(4)}│🟢 متاح")
        lines.append("")
    lines.append("─"*55)
    lines.append(f"  إجمالي البريكات: {len(assigns)}  ·  النظام: 15·15·20·10")
    lines.append(f"  الموظفين: {len(emps)}  ·  {datetime.now().strftime('%I:%M %p')}")
    lines.append("─"*55)
    return "\n".join(lines)

@app.route("/api/send-email", methods=["POST"])
def api_send_email():
    try:
        d = request.json
        smtp = d.get("smtp", {})
        for r in ["host","port","user","pass","from_email"]:
            if r not in smtp or not smtp[r]:
                return jsonify({"error":f"⚠️ SMTP: {r} مطلوب"}), 400

        assigns = d.get("assignments", {})
        shifts = d.get("shifts", {})
        vo = d.get("view_only", {})
        dur = d.get("durations", [15,15,20,10])
        cdur = d.get("custom_durations", {})
        emps = d.get("employees", [])
        extras = d.get("extra_emails", [])

        to_list = list(set([e["email"] for e in emps if e.get("email")] +
                          [e["email"] for e in extras if e.get("email")]))
        if not to_list: return jsonify({"error":"⚠️ لا يوجد مستلمين"}), 400

        body = build_email_body(assigns, shifts, vo, dur, cdur, emps, extras)
        msg = MIMEMultipart("alternative")
        dt = datetime.now().strftime('%Y-%m-%d')
        msg["Subject"] = f"📋 Break Schedule - Al Mana General Hospitals - {dt}"
        msg["From"] = smtp["from_email"]
        msg["To"] = ", ".join(to_list[:50])

        html = f"""<!DOCTYPE html><html dir="rtl"><head><meta charset="UTF-8">
<style>body{{font-family:Tahoma,sans-serif;direction:rtl;padding:20px;background:#f5f5f5}}
pre{{background:#fff;padding:15px;border-radius:10px;border:1px solid #ddd;font-size:13px;line-height:1.6}}
</style></head><body><pre>{body.replace(chr(10),"<br>")}</pre>
<p style="color:#999;font-size:11px;text-align:center">Al Mana General Hospitals · مستشفيات المانع العامة · نظام البريكات</p></body></html>"""
        msg.attach(MIMEText(body, "plain"))
        msg.attach(MIMEText(html, "html"))

        port = int(smtp["port"])
        ctx = ssl.create_default_context()
        if port == 465:
            with smtplib.SMTP_SSL(smtp["host"], port, context=ctx) as s:
                s.login(smtp["user"], smtp["pass"])
                s.sendmail(smtp["from_email"], to_list, msg.as_string())
        else:
            with smtplib.SMTP(smtp["host"], port) as s:
                s.starttls(context=ctx)
                s.login(smtp["user"], smtp["pass"])
                s.sendmail(smtp["from_email"], to_list, msg.as_string())

        return jsonify({"ok":True, "sent":len(to_list), "message":f"✅ تم الإرسال لـ {len(to_list)} شخص"})
    except smtplib.SMTPAuthenticationError:
        return jsonify({"error":"❌ خطأ في بيانات SMTP"}), 401
    except smtplib.SMTPException as e:
        return jsonify({"error":f"❌ SMTP: {str(e)[:100]}"}), 500
    except Exception as e:
        return jsonify({"error":f"❌ {str(e)[:100]}"}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT",5000))
    print(f"Break Schedule Server @ http://localhost:{port}")
    print(f"Admin:   http://localhost:{port}/schedule.html")
    print(f"Employee: http://localhost:{port}/employee.html")
    app.run(host="0.0.0.0", port=port, debug=False)
